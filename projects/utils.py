from .models import *
from django.db.models import Q
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors, fills, PatternFill, Border, Side, Alignment, Font, numbers
from openpyxl.drawing.image import Image
from config import settings
import datetime


def searchProjects(request):
    
    query = Q()
    
    keywords = ''
    if request.GET.get('keywords'):
        keywords = request.GET.get('keywords')
        query.add(Q(title__icontains=keywords),Q.AND)
     
    selected_status = ''
    if request.GET.get('status'):
        selected_status = request.GET.get('status')
        query.add(Q(status__status__icontains=selected_status),Q.AND)
        
    selected_category = ''
    if request.GET.get('category'):
        selected_category = request.GET.get('category')
        query.add(Q(status__category__name__icontains=selected_category),Q.AND)
        
    selected_donor = ''
    if request.GET.get('donor'):
        selected_donor = request.GET.get('donor')
        donor = Donor.objects.filter(name__iexact=selected_donor)
        query.add(Q(donors__in=donor),Q.AND)
        
    selected_implementor = ''
    if request.GET.get('implementor'):
        selected_implementor = request.GET.get('implementor')
        implementor = Implementor.objects.filter(name__iexact=selected_implementor)
        query.add(Q(implementors__in=implementor),Q.AND)
        
    selected_partner = ''
    if request.GET.get('partner'):
        selected_partner = request.GET.get('partner')
        partner = Partner.objects.filter(name__iexact=selected_partner)
        query.add(Q(partners__in=partner),Q.AND)
        
    selected_type = ''
    if request.GET.get('type'):
        selected_type = request.GET.get('type')
        query.add(Q(type__name__icontains=selected_type),Q.AND)
        
    selected_risk = ''
    if request.GET.get('risk'):
        selected_risk = request.GET.get('risk')
        query.add(Q(risk_rate__rate__icontains=selected_risk),Q.AND)
    
    projects = Project.objects.filter(query)
    
    tags = ''
    if request.GET.get('tags'):
        # split tags
        tags = request.GET.get('tags').replace(" ","").split(',')
        # filter first tag
        projects = projects.filter(tags__in=Tag.objects.filter(name__icontains=tags[0]))
        # chain consecutive tags in list
        for tag in tags[1:]:
            projects = projects.filter(tags__in=Tag.objects.filter(name__icontains=tag))
        
        tags = ','.join(str(tag) for tag in tags)  

   
    selected = {
        'keywords' : keywords,
        'tags' : tags,
        'selected_status': selected_status,
        'selected_category': selected_category,
        'selected_donor': selected_donor,
        'selected_implementor': selected_implementor, 
        'selected_partner': selected_partner, 
        'selected_type': selected_type, 
        'selected_risk': selected_risk, 
    }

    return projects, selected

def getProjectFilters():
    search_filters = {
        'donors' :  Donor.objects.all(),
        'implementors': Implementor.objects.all(),
        'partners': Partner.objects.all(),
        'risks': Risk.objects.all(),
        'types': ProjectType.objects.all(),
        'categories': StatusCategory.objects.all(),
        'statuses': Status.objects.all(),
    }
    return search_filters

def getDataForExcel():
    
    # grab queryset
    projects = Project.objects.all()
    
    # traverse queryset, generate dictionary from each object, append to list
    projects_list = []
    for project in projects:
        projects_list.append(
            { 
                'id': str(project.id),
                'title' : str(project.title),
                'description' : str(project.description),
                'risk_rate' : str(project.risk_rate),
                'type' : str(project.type),
                'status' : str(project.status),
                'donors' : ','.join(str(item) for item in [item for item in project.donors.values_list('name', flat=True)]),
                'implementors' : ','.join(str(item) for item in [item for item in project.implementors.values_list('name', flat=True)]),
                'partners' : ','.join(str(item) for item in [item for item in project.partners.values_list('name', flat=True)]),
                'tags' : ','.join(str(item) for item in [item for item in project.tags.values_list('name', flat=True)]),
            }
        )
    
    # return list 
    return projects_list
       
def createRowItemsFromJSON(data):
    
    data = json.loads(data)
    
    col_headers = [
        'Id', 
        'Title',
        'Description',
        'Risk Rate',
        'Type',
        'Status',
        'Donors',
        'Implementing Agencies',
        'Partner Organisations\\Acreddited Entites',
        'Tags'
        ]
    
    project_list = []
    for project in data:
        print(project)
        project_list.append([
            project['pk'],
            project['fields']['title'],
            project['fields']['description'],
            project['fields']['risk_rate'],
            project['fields']['type'],
            project['fields']['status'],
            ', '.join(item for item in project['fields']['donors']),
            ', '.join(item for item in project['fields']['implementors']),
            ', '.join(item for item in project['fields']['partners']),
            ', '.join(item for item in project['fields']['tags']),
        ])
        
    return col_headers, project_list   
        

       
def filterProjectsForReport(data, report_type):
    
    projects = json.loads(data)
    
    # grab donors from each project
    donors_lists = [project['fields']['donors'] for project in projects]  

    # extend all donors list to a donors list so that it becomes a list of strings
    donors = []
    for _list in donors_lists:
        donors.extend(_list)

    # remove duplicate donors
    donors = list(set(donors))
    
    donor_objs = []
    for donor in donors:
        donor_objs.append({
            'name': donor,
            'project_count': 0,
            'total_funding': 0,
        })
    
    
    for project in projects:
        for donor in donor_objs:
            if donor.get('name') in project['fields']['donors']:
                donor['project_count'] = donor.get('project_count') + 1
                if project['fields']['budget_total']:
                    donor['total_funding'] = donor.get('total_funding') + float(project['fields']['budget_total'])
    
    return donor_objs



def generateReport(results, response):
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'

    column_titles = [
        'Donor', 
        'Number of Projects',
        'Total Funding',
        ]
    
    ws.append(column_titles)
    
    for item in results:
        ws.append([
            item.get('name'),
            item.get('project_count'),
            item.get('total_funding')
            ])
    
    ws.append(['','Total',sum([item['total_funding'] for item in results])])
        
    applyReportStyling(column_titles, ws)
    
    # add date row to the bottom, add styles
    date_row = str(ws.max_row+2)
    
    # merge date cells
    ws.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=3)
    
    # append current date and time to date row
    now = datetime.datetime.now()
    formatted_date = now.strftime("%I:%M %p %d-%m-%Y")
    ws['A'+date_row].value = "Report generated at " + formatted_date
    
    # style date row
    ws['A'+date_row].fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color('D9D9D9'))
    ws['A'+date_row].font = Font(name='Arial Narrow',color=colors.Color('1F497D'),size=11,bold=True)
    ws['A'+date_row].alignment = Alignment(horizontal='right')
    
    wb.save(response)




# Styling for Generated Reports
def applyReportStyling(col_headers,ws):
    
    
    bg_color = '4F81BD'
    font_color = colors.WHITE
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    
    default_font = Font(name='Arial Narrow',color=font_color,size=11,bold=False)    
    default_fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(bg_color))
    
    
    
    # Going by each column from left to right
    for col in range(1, len(col_headers) + 1):
        
        # grabbing the column leter
        char = get_column_letter(col)
        
        # setting the column width
        ws.column_dimensions[char].width = 30 
        
        # applying styles to each cell in column
        for cell in ws[char]:
            cell.border = thin_border
            cell.font = default_font
            cell.fill = default_fill
    
    
    # creating the header
    ws.insert_rows(1, amount=4)
    ws.merge_cells("A1:C4")
    ws['A1'].value = 'Generate Report By Donor for filtered Projects'
    ws['A1'].border = thin_border
    ws['A1'].font = Font(name='Arial Narrow',color=font_color,size=11,bold=True) 
    ws['A1'].fill = default_fill
    ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
    
    
    # applying style to titles
    for col in range(1, len(col_headers) + 1):
        titles_row_number = 5
        col_letter = get_column_letter(col)
        ws[col_letter+str(titles_row_number)].font = Font(name='Arial Narrow',color=font_color,size=14,bold=True)
        
    
    # inserting the logo at the top cell
    logo = Image(settings.BASE_DIR / "sig-logo.png")
    logo.width = 50
    logo.height = 50
    ws.add_image(logo, 'A1')
    

    # apply currency format to all cells in 'Total Funding' column
    for row_number in range(6, ws.max_row + 1):
        ws['C'+str(row_number)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # applying style to totals row
    for col in range(1, len(col_headers) + 1):
        col_letter = get_column_letter(col)
        ws[col_letter+str(ws.max_row)].font = Font(name='Arial Narrow',color=font_color,size=16,bold=True)
       
       
    

    
    
    
    
    