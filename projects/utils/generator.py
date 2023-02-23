from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors, fills, PatternFill, Border, Side, Alignment, Font, numbers
from openpyxl.drawing.image import Image
from config import settings
import datetime
import json




def generateSingleSpreadsheet(response,sheet_name, col_headers, row_items):
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws['A1'].value = sheet_name
    last_column_letter = get_column_letter(len(col_headers))
    ws.merge_cells("A1:" + last_column_letter + "1")
    
    ws.append(col_headers)
    for item in row_items:
        ws.append(item)
        
    applyDefaultStyle(ws, col_headers)

    wb.save(response)
    
    
def applyDefaultStyle(ws, col_headers):
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    
    font = Font(color=colors.BLACK,size=11,bold=False)    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='f8f9fa'))
    
    
    for col in range(1, len(col_headers) + 1):
        char = get_column_letter(col)
        ws.column_dimensions[char].width = 20
        for cell in ws[char]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left",vertical="center")
            cell.font = font
            cell.fill = fill
    
    
    title_font = Font(color=colors.WHITE,size=14,bold=True)
    title_fill = PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='212529'))
    header_font =  Font(color=colors.WHITE,size=12,bold=True)
    header_fill = PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='212529'))    
    for col in range(1, len(col_headers) + 1):
        char = get_column_letter(col)
        ws[char+'1'].font = title_font
        ws[char+'1'].fill = title_fill
        ws[char+'2'].font = header_font
        ws[char+'2'].fill = header_fill
        

def generateReport(report_option, results, response):
    
    wb = Workbook()
    
    # Summary Report Sheet
    createSummaryReportSheet(report_option, wb, results)
   
    # Projects by filter Sheet
    createProjectsByFilterSheet(report_option, wb, results)
    
    wb.save(response)



def createProjectsByFilterSheet(report_option, wb, projects):
    
    wb.create_sheet('Projects By ' + report_option)
    ws = wb['Projects By ' + report_option]   
    

    # styles
    header_font = Font(name='Arial Narrow', color=colors.Color('1F497D'),size=15,bold=True)    
    header_fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.WHITE)
    title_font = Font(name='Arial Narrow', color=colors.WHITE,size=14,bold=True)    
    title_fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color('4F81BD'))
    cell_font = Font(name='Arial Narrow', color=colors.BLACK,size=11,bold=False)    
    cell_fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color('DCE6F1'))
    cell_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    align_cell_right = Alignment(horizontal="right",vertical="center")
    align_cell_left = Alignment(horizontal="left",vertical="center")
    
    
    # setting cell width for columns A & B
    for column in range(1,3):
        column = get_column_letter(column)
        ws.column_dimensions[column].width = 30 
    
    
    for project in projects:
        
        # append header row
        ws.append([project.get('filtered_by'),''])
        
        # merge that row
        ws.merge_cells('A' + str(ws.max_row)+':B'+str(ws.max_row))
        
        # style that row
        ws['A'+str(ws.max_row)].font = header_font
        ws['A'+str(ws.max_row)].fill = header_fill
        
        # append title row
        ws.append(['Project','Funding'])

        # style title
        
        ws['A'+str(ws.max_row)].font = title_font
        ws['A'+str(ws.max_row)].fill = title_fill
        
        ws['B'+str(ws.max_row)].font = title_font
        ws['B'+str(ws.max_row)].fill = title_fill
        ws['B'+str(ws.max_row)].alignment = align_cell_right
        
        
        for item in project['projects']:
            project = item['fields']
            
            # append cell row
            ws.append([project.get('title'),float(project.get('budget_total'))])
            
            # style cell row
            
            ws['A'+str(ws.max_row)].font = cell_font
            ws['A'+str(ws.max_row)].fill = cell_fill
            ws['A'+str(ws.max_row)].border = cell_border
            
            ws['B'+str(ws.max_row)].font = cell_font
            ws['B'+str(ws.max_row)].fill = cell_fill
            ws['B'+str(ws.max_row)].border = cell_border
            ws['B'+str(ws.max_row)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            
        # append empty row for spacing   
        ws.append(['',''])
    
    # styleProjectsSheet(ws)
    
    insertDate(ws)



# def styleProjectsSheet(ws):



def createSummaryReportSheet(report_option, wb, results):
    
    ws = wb.active
    ws.title = 'Summary Report'

    column_titles = [
        report_option, 
        'Number of Projects',
        'Total Funding',
        ]
    
    ws.append(column_titles)
    
    for item in results:
        ws.append([
            item.get('filtered_by'),
            item.get('project_count'),
            item.get('total_funding')
            ])
    
    ws.append(['','Total',sum([item['total_funding'] for item in results])])
        
    applyReportStyling(column_titles, ws)
    
    insertDate(ws)
    
    
    
    
    
    
    
def insertDate(ws):
      
    # add date row to the bottom, add styles
    date_row = str(ws.max_row+2)
    
    # merge date cells
    ws.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=3)
    
    # append current date and time to date row
    now = datetime.datetime.now()
    formatted_date = now.strftime("%I:%M %p %d-%m-%Y")
    ws['A'+date_row].value = "Report generated at " + formatted_date
    
    # style date row
    ws['A'+date_row].fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color('D9D9D9'))
    ws['A'+date_row].font = Font(name='Arial Narrow',color=colors.Color('1F497D'),size=11,bold=True)
    ws['A'+date_row].alignment = Alignment(horizontal='right')
    
    
    
    
    
# Styling for Generated Reports
def applyReportStyling(col_headers,ws):
    
    
    bg_color = '4F81BD'
    font_color = colors.WHITE
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    
    default_font = Font(name='Arial Narrow',color=font_color,size=11,bold=False)    
    default_fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(bg_color))
    
    
    
    # Going by each column from left to right
    for col in range(1, len(col_headers) + 1):
        
        # grabbing the column leter
        char = get_column_letter(col)
        
        # setting the column width
        ws.column_dimensions[char].width = 30 
        
        # applying styles to each cell in column
        for cell in ws[char]:
            cell.border = thin_border
            cell.font = default_font
            cell.fill = default_fill
    
    
    # creating the header
    ws.insert_rows(1, amount=4)
    ws.merge_cells("A1:C4")
    ws['A1'].value = 'Summary Report'
    ws['A1'].border = thin_border
    ws['A1'].font = Font(name='Arial Narrow',color=font_color,size=11,bold=True) 
    ws['A1'].fill = default_fill
    ws['A1'].alignment = Alignment(horizontal="center",vertical="center")
    
    
    # applying style to titles
    for col in range(1, len(col_headers) + 1):
        titles_row_number = 5
        col_letter = get_column_letter(col)
        ws[col_letter+str(titles_row_number)].font = Font(name='Arial Narrow',color=font_color,size=14,bold=True)
        
    
    # inserting the logo at the top cell
    logo = Image(settings.BASE_DIR / "sig-logo.png")
    logo.width = 50
    logo.height = 50
    ws.add_image(logo, 'A1')
    

    # apply currency format to all cells in 'Total Funding' column
    for row_number in range(6, ws.max_row + 1):
        ws['C'+str(row_number)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # applying style to totals row
    for col in range(1, len(col_headers) + 1):
        col_letter = get_column_letter(col)
        ws[col_letter+str(ws.max_row)].font = Font(name='Arial Narrow',color=font_color,size=16,bold=True)
        
        
def filterForReportType(report_option, projects):
    
    # handles the report_option 'status'. status key is a list [status, status_category], this grabs the status_category
    if report_option == 'status':
         items = [project['fields'][report_option][1] for project in projects]
    # handles report_option keys 'donors' and 'implementors'   
    else:
        items_lists = [project['fields'][report_option] for project in projects]
        items = []
        for _list in items_lists:
            items.extend(_list)

    # removes duplicate items 
    items = list(set(items))
    
    
    objs = [{'filtered_by': item, 'project_count': 0, 'total_funding': 0, 'projects': []} for item in items]

    
    # loops through each project, get all projects that match filter,
    for project in projects:
        for obj in objs:
            if obj.get('filtered_by') in project['fields'][report_option]:
                obj['projects'].append(project)
                obj['project_count'] = obj.get('project_count') + 1
                if project['fields']['budget_total']:
                    obj['total_funding'] = obj.get('total_funding') + float(project['fields']['budget_total'])
    
    return objs



def createRowItemsFromJson(data):
    
    data = json.loads(data)
    
    col_headers = [
        'Id', 
        'Title',
        'Description',
        'Risk Rate',
        'Type',
        'Status',
        'Donors',
        'Implementing Agencies',
        'Partner Organisations\\Acreddited Entites',
        'Tags'
        ]
    
    project_list = []
    for project in data:
        print(project)
        project_list.append([
            project['pk'],
            project['fields']['title'],
            project['fields']['description'],
            project['fields']['risk_rate'],
            project['fields']['type'],
            project['fields']['status'],
            ', '.join(item for item in project['fields']['donors']),
            ', '.join(item for item in project['fields']['implementors']),
            ', '.join(item for item in project['fields']['partners']),
            ', '.join(item for item in project['fields']['tags']),
        ])
        
    return col_headers, project_list