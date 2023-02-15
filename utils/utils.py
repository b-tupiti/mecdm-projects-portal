from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors, fills, PatternFill, Border, Side, Alignment, Protection, Font, GradientFill


def paginateItems(request, objects, results):

    page = request.GET.get('page')
    paginator = Paginator(objects, results)

    try:
        objects = paginator.page(page)
    except PageNotAnInteger:
        page = 1
        objects = paginator.page(page)
    except EmptyPage:
        page = paginator.num_pages
        objects = paginator.page(page)

    leftIndex = (int(page) - 4)

    if leftIndex < 1:
        leftIndex = 1

    rightIndex = (int(page) + 5)

    if rightIndex > paginator.num_pages:
        rightIndex = paginator.num_pages + 1

    custom_range = range(leftIndex, rightIndex)

    return custom_range, objects



def generateSingleSpreadsheet(response,sheet_name, col_headers, row_items):
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws['A1'].value = sheet_name
    last_column_letter = get_column_letter(len(col_headers))
    ws.merge_cells("A1:" + last_column_letter + "1")
    
    ws.append(col_headers)
    for item in row_items:
        ws.append(item)
        
    applyDefaultStyle(ws, col_headers)

    wb.save(response)
    
    
def applyDefaultStyle(ws, col_headers):
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    
    font = Font(color=colors.BLACK,size=11,bold=False)    
    fill = fills.PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='f8f9fa'))
    
    
    for col in range(1, len(col_headers) + 1):
        char = get_column_letter(col)
        ws.column_dimensions[char].width = 20
        for cell in ws[char]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left",vertical="center")
            cell.font = font
            cell.fill = fill
    
    
    title_font = Font(color=colors.WHITE,size=14,bold=True)
    title_fill = PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='212529'))
    header_font =  Font(color=colors.WHITE,size=12,bold=True)
    header_fill = PatternFill(patternType='solid',
                                  fgColor=colors.Color(rgb='212529'))    
    for col in range(1, len(col_headers) + 1):
        char = get_column_letter(col)
        ws[char+'1'].font = title_font
        ws[char+'1'].fill = title_fill
        ws[char+'2'].font = header_font
        ws[char+'2'].fill = header_fill